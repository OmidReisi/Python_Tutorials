import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

with open(r"./data.json", mode="r") as json_file:
    data = json.load(json_file)

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ["Name", *data["Joe"].keys()]

ws.append(headings)

for person in data:
    grades = data[person].values()
    ws.append([person, *grades])


for col in range(2, len(data["Joe"]) + 2):
    char = get_column_letter(col)
    # we write the excel formula in our string after an equal sign(=)
    ws[char + "7"] = f"=SUM({char + '2'}:{char +'6'})/{len(data)}"

# in openpyxl you can't style a whole row or column and you have to style each cell seperatly.
for col in range(1, 6):
    char = get_column_letter(col)
    ws[char + "1"].font = Font(bold=True, color="0099CCFF")


wb.save(r"./new_grades.xlsx")
