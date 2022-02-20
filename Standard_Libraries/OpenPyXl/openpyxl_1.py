# this module allows us to automate excel worksheets with python (pip install openpyxl)
from openpyxl import load_workbook


# this method loads an existing workbook and returns a workbook object.
# if you set the readonly to true you can't modify the workbook
wb = load_workbook(r"./Centrifugal Pump Handbook_1.xlsx")

# this gives us the current active sheet in the workbook
ws = wb.active
print(ws)

# in order to get a value of a cell we just index it to the sheet.(this returns an object to get the value you have to use .value attribute)
print(ws["C2"].value)


# you can modify a cell by assigning a value to the cell
# when modifying a cell you don't need to use .value attribute, but it's okey if you use it.
ws["H1"].value = "Omid"

# in order for your changes to take effect you need to save the file first.
# remember that you can't save an open file.
# wb.save(r"./Centrifugal Pump Handbook_1.xlsx")

# this gives a list of sheets in the workbook
print(wb.sheetnames)

# you can access different sheets by indexing them to the workbook.
ws = wb["Pump Data"]
print(ws)


# you can create new sheets with this method
# remember that all changes won't take effect until the workbook is saved.
wb.create_sheet("test")
print(wb.sheetnames)
