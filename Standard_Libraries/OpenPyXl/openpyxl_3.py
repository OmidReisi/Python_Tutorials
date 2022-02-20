from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(r"./test.xlsx")
ws = wb.active

# this is how we loop through multiple rows and columns
for row in range(1, 11):
    for col in range(1, 5):

        # this method converts an integer to it's corresponding column('A' -> 1, 'B' -> 2, 'C' -> 3, ...)
        char = get_column_letter(col)

        ws[char + str(row)] = char + str(row)

wb.save(r"./test.xlsx")
