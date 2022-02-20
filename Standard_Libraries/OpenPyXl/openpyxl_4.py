from openpyxl import load_workbook

wb = load_workbook(r"./test.xlsx")
ws = wb.active

ws.merge_cells("E1:H4")

ws.unmerge_cells("E1:H4")

# insert the given amount of rows at the starting index
ws.insert_rows(idx=4, amount=3)
# works just like insert_rows()
ws.delete_rows(idx=4, amount=5)

# you have to pass column number instead of it's letter
ws.insert_cols(idx=3, amount=3)

ws.delete_cols(idx=3, amount=3)

# move a cell range by the given amount of rows and cols(rows>0 -> down, cols>0 -> right)
ws.move_range("A1:B4", rows=5, cols=5)

wb.save(r"./test.xlsx")
